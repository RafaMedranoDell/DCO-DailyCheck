import logging
import smtplib
import socket
import pandas as pd
import numpy as np
from textwrap import indent
from numbers import Number
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from datetime import datetime
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension
import common.functions as fn

# Configure module logger
logger = fn.get_module_logger(__name__)

# Char used to join together HTML code (empty by default)
JOIN_CHAR = ""

# Standarized set of colors to use for styling the tables
RED           = 'background-color: #C70039; color: #a10000'
PASTEL_RED    = 'background-color: #f7b3b3; color: #a10000'
GREEN         = 'background-color: #b3e6b3; color: #2d6a2d'
PASTEL_GREEN  = 'background-color: #b3e6b3; color: #2d6a2d'
YELLOW        = 'background-color: #fff3b3; color: #806600'
PASTEL_YELLOW = 'background-color: #fff2b3; color: #7a7a00'
PASTEL_BLUE   = 'background-color: #85c1e9; color: #ffffff'
PASTEL_ORANGE = 'background-color: #fad5a5; color: #101010'

# Color arrays for rating
COLORS_GYOR = [PASTEL_GREEN, PASTEL_YELLOW, PASTEL_ORANGE, PASTEL_RED]
COLORS_ROYG = list(reversed(COLORS_GYOR))
COLORS_GYR = [PASTEL_GREEN, PASTEL_YELLOW, PASTEL_RED]
COLORS_RYG = list(reversed(COLORS_GYR))

# Centralized style definitions
HEADER_STYLES = {
    'background-color': 'DodgerBlue',
    'color': 'white',
    'font-family': 'Arial, sans-serif',
    'font-size': '11px',
    'font-weight': 'bold',
    'padding': '2px 4px',
    'border': '1px solid #ddd',
    'line-height': '1.2',
    'height': '18px'
}
CELL_STYLES = {
    'font-family': 'Arial, sans-serif',
    'font-size': '11px',
    'padding': '2px 4px',
    'border': '1px solid #ddd',
    'line-height': '1.2',
    'height': '16px'
}
TABLE_STYLES = {
    'border-collapse': 'collapse',
    'width': '100%',
    'font-family': 'Arial, sans-serif',
    'font-size': '11px',
    'mso-table-lspace': '0pt',
    'mso-table-rspace': '0pt',
    'border': '1px solid #ddd'
}
FIXED_TABLE_STYLES = {**TABLE_STYLES, 'table-layout': 'fixed'}

# Minimal table_style for non-table elements
table_style = """
<style>
    .table-cell {
        padding: 5px !important;
        text-align: left;
        vertical-align: top !important;
    }
    .tableset td {
        vertical-align: top !important;
    }
    .tableset table {
        width: 100%;
    }
</style>
"""

# HTML constants
report_begin = f"""
<html xmlns:v="urn:schemas-microsoft-com:vml" xmlns:o="urn:schemas-microsoft-com:office:office" xmlns:w="urn:schemas-microsoft-com:office:word" xmlns:m="http://schemas.microsoft.com/office/2004/12/omml" xmlns="http://www.w3.org/TR/REC-html40">
    <head>
        {table_style}
        <!--[if gte mso 9]>
        <xml>
            <o:OfficeDocumentSettings>
                <o:AllowPNG/>
                <o:PixelsPerInch>96</o:PixelsPerInch>
            </o:OfficeDocumentSettings>
        </xml>
        <![endif]-->
    </head>
"""

report_body = """
    <body style="margin: 0; padding: 0; font-family: Arial, sans-serif;">
        <div style="width: 100%; max-width: 1200px; margin: 0 auto;">
            <h1 style="font-family: Arial, sans-serif; color: #0044cc; margin-bottom: 10px;">{report_title}</h1>
"""

report_end = """
        </div>
    </body>
</html>
"""

report_header2_begin = """
<h2 id="{index_id}" style="font-family: Arial, sans-serif; color: #0066cc; margin: 10px 0;">{header2}</h2>
"""

report_header2_end = ""

report_header3_begin = """
<h3 id="{index_id}" style="font-family: Arial, sans-serif; color: #0066cc; margin: 10px 0;">{header3}</h3>
"""

report_header3_end = ""

report_header4_begin = """
    <h4 id="{index_id}" style="font-family: Arial, sans-serif; color: #0066cc; margin: 10px 0;">{header4}</h4>
"""

report_header4_body = ""

report_header4_end = ""


report_tableset_begin = """
    <table class="tableset" cellpadding="0" cellspacing="0" border="0" width="100%" style="table-layout: fixed;">
        <tr>
"""

report_tableset_column_begin = """
        <td>
        <table cellpadding="0" cellspacing="0" border="0";>
"""
report_tableset_column_item_begin = """
            <tr>
"""

report_tableset_column_item_end = """
            </tr>
"""

report_tableset_column_end = """
        </table>
        </td>
"""
report_tableset_end = """
        </tr>
    </table>
"""

report_table_begin = """
                <td class="table-cell">
"""

report_table_body = """
                    <h5 style="font-family: Arial, sans-serif; color: #0066cc; margin: 10px 0;">
                        {title}
                    </h5>
                        {html_code}
                    <hr style="border: 0; height: 1px; background: #fff; margin: 10px 0;">
"""

report_table_end = """
                </td>
"""

##################################################################
# Styler helper functions
##################################################################

def csv_to_styleddf(system, instance, data_type, dcocfg, fixed=False):
    """
    Loads a CSV file from a system/instance/data_type into a dataframe and applies it the default styler.
    Uses a DCOconfig object

    Parameters:
    - system (str): type of product
    - instance (str): individual ocurrence of the system
    - data_type (str): keyword in the configuration describing the contents
    - dcocfg (DCOconfig): DCOconfig object with the configuration loaded
    - fixed (bool): format the table with equal-width columns

    Returns:
    - Dataframe styled (pd.io.formats.style.Styler) with the contents of the CSV file
    """
    df = dcocfg.load_csv_to_dataframe(system, instance, data_type)
    return df.style if df.empty else table_base_styler(df, fixed)

def table_base_styler(df, fixed=False):
    """
    Applies default table styles:
    - Replaces NaN with empty string
    - Hides index
    - Applies inline styles: first column left-aligned, others centered, headers #cccccc
    - Alternating row colors

    Parameters
    - df (Pandas DataFrame): table to style
    - fixed (bool): equal-width columns with table-layout: fixed

    Returns:
    - Styled Dataframe (pd.io.formats.style.Styler)
    """
    df = df.fillna('')

    # Set table attributes
    table_attrs = FIXED_TABLE_STYLES if fixed else TABLE_STYLES
    table_attrs_str = '; '.join(f'{k}: {v}' for k, v in table_attrs.items())
    df_styler = df.style.set_table_attributes(f'style="{table_attrs_str}"').hide(axis='index')

    # Apply header styles
    header_styles = [
        {'selector': 'th', 'props': [(k, v) for k, v in HEADER_STYLES.items()] + [('text-align', 'center')]},
        {'selector': 'th.col0', 'props': [('text-align', 'left')]}
    ]
    df_styler = df_styler.set_table_styles(header_styles)

    # Apply cell styles
    for i, col in enumerate(df.columns):
        align = 'left' if i == 0 else 'center'
        df_styler = df_styler.set_properties(**{**CELL_STYLES, 'text-align': align}, subset=[col])

    # Apply alternating row colors
    def highlight_rows(row):
        return ['background-color: #ebf5fb' if row.name % 2 == 1 else '' for _ in row]
    df_styler = df_styler.apply(highlight_rows, axis=1)

    return df_styler


def column_wordwrap(styled_df, columns):
    """
    Applies left-alignment and word-wrapping to specified columns.

    Parameters:
    - styled_df (pd.io.formats.style.Styler): Styled Dataframe
    - columns (list of str): list of the column names

    Returns:
    - Styled Dataframe (pd.io.formats.style.Styler)
    """
    return styled_df.set_properties(**{
        'text-align': 'left',
        'white-space': 'normal',
        'word-wrap': 'break-word'
    }, subset=columns)

"""
    Helper functions rate_num_value, rate_num_rows, key_color_value and key_color_rows can be used in two differen ways:
    - Defining a coloring function using functools.partial
        colorByStatus = functools.partial(
            DCOreport.key_color_rows,
            column="Status",
            key_color={"Warning": DCOreport.YELLOW, "Critical": DCOreport.PASTEL_RED})
        df_styler.apply(colorByStatus, axis=1)

    - Direct use in an pandas.io.formats.style.Styler.apply:
        df_styler.apply(
            DCOreport.key_color_rows,
            axis=1,
            column="Status", key_color={"Warning": DCOreport.YELLOW, "Critical": DCOreport.PASTEL_RED})
"""


def format_nums_by_rowid(styleddf, row_label, fmt):
    """
    Apply a numeric format to all cells in a specific row of a Pandas Styler object,
    including the first column if applicable.

    Parameters
    ----------
    styleddf : pandas.io.formats.style.Styler
        The Styler object to apply the formatting to.
    row_label : str
        The value in the first column that identifies the target row.
    fmt : str
        A format string compatible with Python's str.format(), e.g.:
        '{:.0f}' for integers without decimals,
        '{:,.2f}' for two decimals with thousands separator.

    Behavior
    --------
    - Locates the row where the first column matches `row_label`.
    - Applies the format `fmt` to all cells in that row (except the cell that equals `row_label`).
    - Formatting affects only the rendered HTML/Excel output via Styler, not the underlying DataFrame values.

    Returns
    -------
    pandas.io.formats.style.Styler
        The Styler object with the format applied, ready for `.to_html()` or `.to_excel()` export.

    Example
    -------
    csCapacity = format_nums_by_rowid(csCapacity, "Used capacity (GB)", "{:.0f}")
    html_output = csCapacity.to_html()
    """

    def fmt_func(val):
        if val == row_label:
            return val  # Keep the label unchanged
        try:
            return fmt.format(float(val))  # Convert to float and format
        except (ValueError, TypeError):
            return val  # If conversion fails, return original value

    # Select all columns of the row where the first column equals row_label
    return styleddf.format(fmt_func, subset=pd.IndexSlice[styleddf.data.iloc[:, 0] == row_label, :])


def rate_num_value(val, rate_intervals, rating, force_conversion=False):
    """
    Maps a numerical value to a rating based on predefined intervals.
    Applies to just one cell.
    Used to colorize numeric values into categories/rates with colors.

    Parameters:
    - rate_intervals (list): Sorted list of numbers defining half-open intervals
        [low, high). The last interval includes the upper bound (val == rate_intervals[-1]).
    - rating (list): List of rating labels, with len(rating) == len(rate_intervals) - 1.
    - val (int or float): Value to map to a rating.
    - force_conversion (bool): force strings to be converted into numbers

    Returns:
    - str: Rating label for the interval containing `val`, or "n/a" if `val` is outside
            the bounds of `rate_intervals`.

    Raises:
        ValueError: If inputs are invalid (e.g., unsorted intervals, mismatched lengths,
            non-numeric val, or insufficient interval boundaries).
        TypeError: If rate_intervals or rating is not a list, or val is not a number.

    Example:
        colorByPercent = functools.partial(
            DCOreport.rate_num_value,
            rate_intervals=[0, 70, 85, 100],
            rating=DCOreport.COLORS_GYR,
            force_conversion=True
        )
        # Colorize column "percent" with the function
        styled_df = DCOreport.apply_styler_map(styled_df, colorByPercent, subset=["percent"])
    """
    # Validate input types
    if not isinstance(rate_intervals, list) or not isinstance(rating, list):
        raise TypeError("rate_intervals and rating must be lists")
    if not isinstance(val, (Number, np.number)):
        if force_conversion:
            try:
                val = int(val) if val.isdigit() else float(val)
            except (ValueError, TypeError):
                logger.warning(f'rate_num_value: Error while trying to convert "{val} of type {type(val)}"')
                return ''
        else:
            logger.info(f'rate_num_value: Not converting "{val}" of type {type(val)}')
            return ''

    # Validate list lengths and minimum requirements
    if len(rate_intervals) < 2:
        raise ValueError("rate_intervals must contain at least two elements")
    if len(rating) != len(rate_intervals) - 1:
        raise ValueError("len(rating) must equal len(rate_intervals) - 1")

    # Check if rate_intervals is sorted
    if not all(rate_intervals[i] <= rate_intervals[i + 1] for i in range(len(rate_intervals) - 1)):
        raise ValueError("rate_intervals must be sorted in ascending order")

    # Check if val is outside the range of intervals
    if val < rate_intervals[0] or val > rate_intervals[-1]:
        return ''

    # Iterate over consecutive interval pairs and corresponding ratings
    for low, high, rate in zip(rate_intervals[:-1], rate_intervals[1:], rating):
        if low <= val < high:
            return rate

    # Return last rating if val equals the final boundary
    return rating[-1]

def rate_num_rows(row, column, rate_intervals, rating, force_conversion=False):
    """
    Applies rating color to entire row based on value in specified column, excluding first column.

    Parameters:
    - row (Pandas series):
    - column (str):
    - rate_intervals (list): Sorted list of numbers defining half-open intervals
        [low, high). The last interval includes the upper bound (val == rate_intervals[-1]).
    - rating (list): List of rating labels, with len(rating) == len(rate_intervals) - 1.
    - force_conversion (bool): force strings to be converted into numbers

    Example:
        colorTempRows = functools.partial(
            DCOreport.rate_num_rows,
            column="Temp Cº",
            rate_intervals=[0, 30, 40, 100],
            rating=DCOreport.COLORS_GYR,
            force_conversion=True
        )
        environmentStatus = environmentStatus.apply(colorTempRows, axis=1)
    """
    color = rate_num_value(row[column], rate_intervals, rating, force_conversion)
    return [color] * len(row)


def key_color_value(val, key_color, def_color=''):
    """
    Maps a key to a rating color, if val not found in key_color dictionary returns default color


    Example:
        colorBySeverityVal = functools.partial(
            DCOreport.key_color_value,
            key_color={
                "Critical": DCOreport.PASTEL_RED,
                "Major": DCOreport.PASTEL_ORANGE,
                "Warning": DCOreport.YELLOW,
                "OK": DCOreport.GREEN}
        )
        systemSummary = DCOreport.apply_styler_map(systemSummary, colorBySeverityVal)
    """
    return key_color.get(val, def_color)

def key_color_rows(row, column, key_color, def_color=''):
    """
    Applies key-based rating color to entire row.

    Example:
        colorBySeverityRow = functools.partial(
            DCOreport.key_color_rows,
            column="Severity",
            key_color={"critical": DCOreport.PASTEL_RED, "warning": DCOreport.PASTEL_YELLOW}
        )
        alertDetail = alertDetail.apply(colorBySeverityRow, axis=1)
    """
    color = key_color_value(row[column], key_color, def_color)
    return [color] * len(row)

def format_cells_by_rowid(styler, row_id, formatter_func):
    """
    Apply a CSS formatter function to cells in a specific row (excluding the first column).

    Parameters:
    - styler (pandas.io.formats.style.Styler):  styled DataFrame
    - row_id (str): value in the first column identifying the target row
    - formatter_func: function, takes a cell value and returns a CSS style string

    Returns:
    - Styled Dataframe (pd.io.formats.style.Styler)
    """
    # Get the name of the first column
    first_col = styler.data.columns[0]

    # Identify the row index where the first column matches row_id
    row_idx = styler.data[styler.data[first_col] == row_id].index

    if not row_idx.empty:
        # Get columns excluding the first column
        cols_to_format = [col for col in styler.data.columns if col != first_col]
        # Apply formatter function to the specified row and columns using the bridge function
        styler = apply_styler_map(styler, formatter_func, subset=(row_idx, cols_to_format))

    return styler

def apply_styler_map(styler, func, subset=None):
    """
    Applies a mapping function to a Styler object, ensuring compatibility
    between Pandas versions (< 2.1.0 uses applymap, >= 2.1.0 uses map).

    Parameters:
    - styler (pd.io.formats.style.Styler): Styled Dataframe
    - func (callable): Function to apply
    - subset (label-like, optional): Slice of entries to apply function to

    Returns:
    - Styled Dataframe (pd.io.formats.style.Styler)
    """
    if hasattr(styler, 'map'):
        return styler.map(func, subset=subset)
    else:
        return styler.applymap(func, subset=subset)


def format_by_rowid(styler, rows_formater):
    """
    Applies one CSS formatter function to the cells in a row (excluding the first column).
    Function is selected by matching the first column with the row_ids provided.
    The rows that doesn't match any provided row_id are not mofidied.

    Parameters:
    - styler (pandas.io.formats.style.Styler):  styled DataFrame
    - rows_formater (list): list of tuples with row_id and formatter_func to apply

    Returns:
    - Styled Dataframe (pd.io.formats.style.Styler)

    Example:
        alertSummary = DCOreport.format_by_rowid(
            alertSummary,
            [("Critical", colorRedNonZero),
            ("Major", colorRedNonZero),
            ("Warning", colorYellowNonZero)]
        )
    """
    for row_id, formatter_func in rows_formater:
        styler = format_cells_by_rowid(styler, row_id, formatter_func)
    return styler

##################################################################
# DCOreport class
##################################################################

class DCOreport:
    """Generates and manages HTML reports for daily checks.

    Creates structured HTML snd Excel reports with styled tables (DodgerBlue headers, first column
    left-aligned, others centered) for browser and email (e.g., Outlook) compatibility.
    Supports hierarchical sections (header2, header3, header4, tableset, table) and email
    delivery via SMTP.

    Methods:
        add_header2(header2): Adds a top-level section header (e.g., "Protection").
        add_header3(header2, header3): Adds a sub-section header under header2 (e.g., "Data Domain").
        add_header4(header2, header3, header4): Defines a header4 for grouping tablesets.
        add_tableset(header2, header3, header4, tableset): Add a group of tables side by side
        add_table(header2, header3, header4, tablename, table, tableset=""): Add a dataframe table to the report
        generate_html(): Returns the complete HTML report as a string.
        save_html(report_name): Saves the report to an HTML file.
        save_xls(self, report_name, multi_sheet=False, indent=True, gen_index=False): Saves the report to an XLSX file.
        send_email(customerName, sender_email, receiver_email, smtp_server, smtp_port,
                   ssl=False, password=None): Sends the report via email using SMTP.
    """
    def __init__(self, title="DCO Daily Check Report"):
        self.root = {"items": {}}
        self.report_title = title

    def _add_item(self, *args, parent=None):
        """
        Generic add an element to the report tree.
        Calls itself recursively to create the branches/leaves.
        """
        if not parent:
            parent = self.root

        item, *remaining = args
        if item not in parent["items"]:
            parent["items"][item] = {"items": {}}
        if remaining:
            self._add_item(*remaining, parent=parent["items"][item])

    def _add_keys(self, *args, key, value, parent=None):
        """
        Add additional properties/keys to the report tree to store key/value pairs.
        Calls itself recursively to locate the branch.
        """
        if not parent:
            parent = self.root

        item, *remaining = args
        if remaining:
            self._add_keys(*remaining, key=key, value=value, parent=parent["items"][item])
        else:
            parent["items"][item][key] = value

    def add_header2(self, header2):
        self._add_item(header2)

    def add_header3(self, header2, header3):
        self._add_item(header2, header3)

    def add_header4(self, header2, header3, header4):
        self._add_item(header2, header3, header4)

    def add_tableset(self, header2, header3, header4, tableset):
        self._add_item(header2, header3, header4, tableset)

    def add_table(self, header2, header3, header4, tablename, table, tableset=""):
        """
        Parameters:
        - header2 (str): header level 2
        - header3 (str): header level 3
        - header4 (str): header level 4
        - tablename (str): title of the table
        - table (pd.io.formats.style.Styler): Styled Dataframe
        - tableset (str): horizontal grouping of one or more tables
        """
        if not isinstance(table, pd.io.formats.style.Styler):
            raise TypeError(f"add_table only supports Pandas Styled tables. Provided {type(table)}")

        tableset, _, tableset_column = tableset.partition("/")
        self._add_item(header2, header3, header4, tableset, tableset_column, tablename)
        self._add_keys(header2, header3, header4, tableset, tableset_column, tablename, key="dataframe", value=table)

    def generate_html(self, gen_index=False):
        """
        Generates the HTML code by walking the report tree.
        Uses the constants for the different sections: header2, header3...
        For each level adds <level>_begin, <level>_body, <level>_end
        Uses SchemaNumbering object to generate an index prefix
        """
        def index_text(text, level):
            return index.str(level) + text if gen_index and text else text

        report_parts = []
        index = SchemaNumbering(4, endstr='. ')
        html_index = HTMLschema()

        # Process header2 level
        for header2_name, header2_data in self.root["items"].items():
            # Skip top level (header2) if it is empty
            if not header2_data["items"]:
                continue
            section_name = index_text(header2_name, 1)
            index_id = html_index.add_item(section_name, 0)
            report_parts.append(report_header2_begin.format(index_id=index_id, header2=section_name))

            # Process header3 level
            for header3_name, header3_data in header2_data["items"].items():
                section_name = index_text(header3_name, 2)
                index_id = html_index.add_item(section_name, 1)
                report_parts.append(report_header3_begin.format(index_id=index_id, header3=section_name))

                # Process header4 level
                for header4_name, header4_data in header3_data["items"].items():
                    section_name = index_text(header4_name, 3)
                    index_id = html_index.add_item(section_name, 2)
                    report_parts.append(report_header4_begin.format(index_id=index_id, header4=section_name))
                    report_parts.append(report_header4_body)

                    # Process tableset level
                    for tableset_name, tableset_data in header4_data["items"].items():
                        report_parts.append(report_tableset_begin)

                        # Process tableset_column level
                        for tableset_column_name, tableset_column_data in tableset_data["items"].items():
                            if tableset_column_name:
                                report_parts.append(report_tableset_column_begin)

                            # Process tables level
                            for table_name, table_data in tableset_column_data["items"].items():
                                if tableset_column_name:
                                    report_parts.append(report_tableset_column_item_begin)

                                report_parts.append(report_table_begin)
                                # Generate the table from the dataframe and indent the html for easier review
                                html_table = table_data["dataframe"].to_html()
                                report_parts.append(report_table_body.format(
                                        title=index_text(table_name, 4),
                                        html_code=indent(html_table, prefix=' '*4*6))
                                )
                                report_parts.append(report_table_end)
                                if tableset_column_name:
                                    report_parts.append(report_tableset_column_item_end)
                                index.levelUp(4)
                            # End of tables level
                            if tableset_column_name:
                                report_parts.append(report_tableset_column_end)
                        # End of tableset_column level
                        report_parts.append(report_tableset_end)
                    # End of tableset level
                    report_parts.append(report_header4_end)
                    index.levelUp(3)
                # End of header4 level
                report_parts.append(report_header3_end)
                index.levelUp(2)
            # End of header3 level
            report_parts.append(report_header2_end)
            index.levelUp(1)
        # End of header2 level
        report_parts.append(report_end)
        return JOIN_CHAR.join([
            report_begin,
            report_body.format(report_title=self.report_title),
            html_index.generate(),
            *report_parts])

    def send_email(self, subject, sender_email, receiver_email, smtp_server, smtp_port, attach_fname=None, ssl=False, password=None):
        if ssl and not password:
            raise ValueError("When using SSL a password must provided")

        # Prepare the email message
        message = MIMEMultipart("alternative")
        message["Subject"] = subject
        message["From"] = sender_email
        message["To"] = receiver_email

        # Attach HTML content as email body
        html_content = self.generate_html()
        html_part = MIMEText(html_content, "html")
        message.attach(html_part)

        if attach_fname:
            # Attach HTML content as a file
            html_attachment = MIMEText(html_content, "html")
            html_attachment.add_header("Content-Disposition", "attachment", filename=attach_fname)
            message.attach(html_attachment)
            logger.debug(f'Attached report: {attach_fname}')

        # Send the email via the specified SMTP server
        try:
            with smtplib.SMTP(smtp_server, smtp_port) as smtp_server:
                if ssl:
                    # Start SSL/TLS communication and login into the server with the provided password
                    smtp_server.starttls()
                    smtp_server.login(sender_email, password)
                smtp_server.sendmail(sender_email, receiver_email, message.as_string())
            logger.info("Email sent successfully")
        except smtplib.SMTPAuthenticationError as e:
            logger.error(f"Authentication Error: Check your sender_email and app_password. Details: {e}")
        except smtplib.SMTPConnectError as e:
            logger.error(f"Connection Error: Could not connect to the SMTP server. Check your internet connection or firewall. Details: {e}")
        except Exception as e:
            logger.error(f"An unexpected error occurred: {e}", exc_info=True)

    def save_html(self, report_name, gen_index=False):
        with open(report_name, "wt") as f:
            f.write(self.generate_html(gen_index=gen_index))
        logger.info(f'Report saved as: {report_name}')

    def save_xls(self, report_name, multi_sheet=False, indent=True, gen_index=False):
        def index_text(text, level):
            return index.str(level) + text if gen_index and text else text

        xlsreport = xlshelper(report_name)

        # Columns for elements when indent=true/false
        if indent:
            headers_indent = {'h2': 1, 'h3': 2, 'h4': 3, 'tableset': 4}
        else:
            headers_indent = {'h2': 0, 'h3': 0, 'h4': 0, 'tableset': 0}

        index = SchemaNumbering(4, endstr='. ')

        # If not multi sheet, the single sheet has the date as name
        if not multi_sheet:
            sheet_name = "Report " + datetime.now().strftime('%Y-%m-%d')
            xlsreport.addSheet(sheet_name)
            xlsreport.writeCell(sheet_name, self.report_title, Font(bold=True, size=30))

        # Process header2 level
        for header2_name, header2_data in self.root["items"].items():
            # Skip top level (header2) if it is empty
            if not header2_data["items"]:
                continue

            # If multi sheet, every header2 has its own sheet
            if multi_sheet:
                sheet_name = header2_name
                xlsreport.addSheet(sheet_name)
                xlsreport.writeCell(sheet_name, self.report_title, Font(bold=True, size=30))
            xlsreport.setCol(sheet_name, headers_indent["h2"])
            if header2_name:
                xlsreport.writeCell(sheet_name, index_text(header2_name, 1), Font(bold=True, size=20))

            # Process header3 level
            for header3_name, header3_data in header2_data["items"].items():
                xlsreport.setCol(sheet_name, headers_indent["h3"])
                if header3_name:
                    xlsreport.writeCell(sheet_name, index_text(header3_name, 2), Font(bold=True, size=16))

                # Process header4 level
                for header4_name, header4_data in header3_data["items"].items():
                    xlsreport.setCol(sheet_name, headers_indent["h4"])
                    if header4_name:
                        xlsreport.writeCell(sheet_name, index_text(header4_name, 3), Font(bold=True, size=12))

                    # Process tableset level
                    for tableset_name, tableset_data in header4_data["items"].items():
                        xlsreport.setCol(sheet_name, headers_indent["tableset"])
                        max_table_rows = 0 # Max rows in the tables of the tableset
                        tbl_title_rows = 0 # Any tables in the tableset had title?

                        # Process tables level
                        for table_name, table_data in tableset_data["items"].items():
                            rows_cnt = len(table_data["dataframe"].data) + 1  # Add the header row
                            cols_cnt = len(table_data["dataframe"].data.columns)
                            if table_name:
                                xlsreport.writeCell(sheet_name, index_text(table_name, 4), Font(bold=True, size=10))
                            xlsreport.writeTable(sheet_name, table_data["dataframe"])
                            # If the table had a title, move the cursor up for the next table
                            if table_name:
                                xlsreport.moveUp(sheet_name, 2)
                                tbl_title_rows = 1
                            # Get the max number of rows the tables have in this tableset
                            max_table_rows = max(max_table_rows, rows_cnt)
                            #location = '/'.join([header2_name, header3_name, header4_name, tableset_name, table_name])
                            #logger.debug(f'{location}: {cols_cnt}:{rows_cnt} [{max_table_rows}]')

                            # Move the cursor right the number of columns the dataframe had plus one empty column
                            xlsreport.moveRight(sheet_name, cols_cnt+1)

                            index.levelUp(4)

                        # After finishing the tables in a tableset, move the cursor down:
                        # table title row +  blank row after the table
                        xlsreport.moveDown(sheet_name, max_table_rows + tbl_title_rows + 2)
                        # End of tables level
                    index.levelUp(3)
                    # End of tableset level
                index.levelUp(2)
                # End of header4  level
            index.levelUp(1)
            # End of header3  level
        # End of header2 level
        xlsreport.close()
        logger.info(f'Report saved as: {report_name}')

class xlshelper:
    """
    Class to manage creating and adding text and dataframes to an excel file.
    For each sheet it maintains a "cursor" to know where to write it next.
    Stores the location of the added dataframes to autoadjust the columns width before saving.
    """
    def __init__(self, report_name):
        self.xlsreport = pd.ExcelWriter(report_name, engine="openpyxl")
        self.cursors = {}
        self.workbook = self.xlsreport.book
        self.table_locations = {}

    def addSheet(self, sheet_name):
        if sheet_name not in self.cursors:
            self.workbook.create_sheet(sheet_name)
            self.cursors[sheet_name] = {}
            self.cursors[sheet_name]["col"] = 0
            self.cursors[sheet_name]["row"] = 0
            self.table_locations[sheet_name] = []

    def _addTableRanges(self, sheet_name, cols, rows):
        self.table_locations[sheet_name].append((
            self.cursors[sheet_name]["col"],
            self.cursors[sheet_name]["row"],
            cols,
            rows
        ))
        #logger.debug(f'{self.cursors[sheet_name]["col"]}, {self.cursors[sheet_name]["row"]}, {cols}, {rows}')

    def moveDown(self, sheet_name, row_delta=1):
        self.cursors[sheet_name]["row"] += row_delta

    def moveUp(self, sheet_name, row_delta=1):
        self.cursors[sheet_name]["row"] -= row_delta

    def moveRight(self, sheet_name, col_delta=1):
        self.cursors[sheet_name]["col"] += col_delta

    def setCol(self, sheet_name, value=0):
        self.cursors[sheet_name]["col"] = value

    def writeCell(self, sheet_name, text, fmt):
        self.addSheet(sheet_name)
        column = self.cursors[sheet_name]["col"]
        row = self.cursors[sheet_name]["row"]
        if isinstance(column, int):
            column = get_column_letter(column+1)

        # worksheet cells start in 1 while DataFrame.to_xls() start in 0. Add 1 to keep them aligned
        cell_address = f'{column}{row+1}'
        self.workbook[sheet_name][cell_address] = text
        self.workbook[sheet_name][cell_address].font = fmt
        self.moveDown(sheet_name, 2)

    def writeTable(self, sheet_name, table):
        if not isinstance(table, pd.io.formats.style.Styler):
            raise TypeError(f"writeTable only supports Pandas Styled tables. Provided {type(table)}")
        self.addSheet(sheet_name)
        tbl_startcol = self.cursors[sheet_name]["col"]
        tbl_startrow = self.cursors[sheet_name]["row"]
        table.to_excel(self.xlsreport, index=False, sheet_name=sheet_name, startcol=tbl_startcol, startrow=tbl_startrow)
        rows_cnt = len(table.data) + 1 # Add the header row
        cols_cnt = len(table.data.columns)
        self._addTableRanges(sheet_name, cols_cnt, rows_cnt)

    def _adjustColumnWidths(self):
        max_width = 50
        for sheet_name in self.table_locations.keys():
            column_max_lengths = {}
            # Iterate through each table location
            for start_col, start_row, col_size, row_size in self.table_locations[sheet_name]:
                start_col += 1
                start_row += 1
                # Process each column in the table
                for col_idx in range(start_col, start_col + col_size):
                    column_letter = get_column_letter(col_idx)

                    # Get max length for cells in this column across all rows in the table
                    max_length = 0
                    for row_idx in range(start_row, start_row + row_size):
                        cell = self.workbook[sheet_name].cell(row=row_idx, column=col_idx)
                        try:
                            # Convert cell value to string and get its length
                            cell_value = str(cell.value) if cell.value is not None else ""
                            cell_length = len(cell_value)
                            max_length = max(max_length, cell_length)
                        except:
                            pass

                    # Update max length for this column if it's greater than previous
                    if column_letter in column_max_lengths:
                        column_max_lengths[column_letter] = max(column_max_lengths[column_letter], max_length)
                    else:
                        column_max_lengths[column_letter] = max_length

            # Set column widths
            for column_letter, max_length in column_max_lengths.items():
                # Add some padding (1.2 factor) and cap at max_width
                adjusted_width = min(max_length * 1.2, max_width)
                # Minimum width to prevent too narrow columns
                adjusted_width = max(adjusted_width, 8.43)  # Default Excel column width
                self.workbook[sheet_name].column_dimensions[column_letter].width = adjusted_width
                #logger.debug(f'Adjust {sheet_name}:{column_letter} -> {adjusted_width}')

    def close(self):
        self._adjustColumnWidths()
        self.xlsreport.close()

class SchemaNumbering:
    def __init__(self, nlevels, endstr=' '):
        self.codes = [1] * nlevels
        self.endstr = endstr
    def levelUp(self, level):
        """
        Sums one to the index number of the provided level, and resets the index numbers in the sublevels to 1
        1.4.3.2 -> levelUp(2) -> 1.5.1.1
        """
        self.codes[level-1] += 1
        for i in range(level, len(self.codes)):
            self.codes[i]=1
    def str(self, nlevels):
        return '.'.join(str(self.codes[x]) for x in range(0, nlevels)) + self.endstr

class HTMLschema:
    def __init__(self):
        self.schema = []
        self._last_div_id = 0

    def add_item(self, text, level):
        self._last_div_id += 1
        indent = level * 2
        div_id = f'schema_{self._last_div_id}'
        self.schema.append(f'<div style="margin-left: {indent}em;"><a href="#{div_id}">{text}</a></div>')

        return div_id

    def generate(self):
        return '\n'.join(['<br>', *self.schema, '<br>', '<hr>','<br>'])